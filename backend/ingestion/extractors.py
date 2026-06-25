"""
Format-specific document extractors for the ingestion pipeline.
Each extractor handles a specific file format and returns a normalised dict.
"""
import csv
import io
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

from backend.observability.logger import get_logger
from backend.observability.metrics import measure_latency

logger = get_logger("extractors")


def _parse_date(date_str: str) -> Optional[datetime]:
    """Try multiple date formats."""
    formats = [
        "%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y", "%d-%m-%Y",
        "%b %d, %Y", "%d %b %Y", "%Y%m%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except (ValueError, AttributeError):
            continue
    return None


def _parse_float(val) -> float:
    """Parse a number from various formats."""
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        cleaned = val.upper().replace("O", "0")
        cleaned = cleaned.replace(",", "").replace("₹", "").replace("INR", "").replace("USD", "").strip()
        if cleaned.count(".") > 1:
            parts = cleaned.rsplit(".", 1)
            cleaned = parts[0].replace(".", "") + "." + parts[1]
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    return 0.0


# ── PDF Extractor ────────────────────────────────────────────────────────────

def extract_pdf(file_path: str) -> dict:
    """Extract structured data from a PDF quotation using pdfplumber."""
    with measure_latency("ingestion", "extract_pdf", file=file_path):
        try:
            import pdfplumber

            text = ""
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"

            if not text.strip():
                logger.warning(f"No text extracted from PDF: {file_path}")
                return {"error": "No text could be extracted", "raw_text": ""}

            data = _extract_from_text(text, "pdf")

            # Fallback to LLM if any important fields are missing
            _important_missing = (
                not data.get("line_items")
                or not data.get("supplier_name")
                or not data.get("warranty")
                or not data.get("freight_terms")
                or not data.get("total_incl_tax")
                or not data.get("delivery_days")
            )
            if _important_missing:
                logger.info(f"Rule-based extraction incomplete, falling back to LLM for {file_path}")
                try:
                    from backend.intelligence.llm_client import extract_quotation_data
                    llm_data = extract_quotation_data(text)
                    if llm_data:
                        # Merge: LLM fills in missing fields, rule-based values take priority
                        for key, val in llm_data.items():
                            if key == "raw_text":
                                continue
                            existing = data.get(key)
                            # Only override if rule-based result is empty/None/zero
                            if not existing or existing == 0:
                                data[key] = val
                        data["raw_text"] = text
                        return data
                except Exception as llm_err:
                    logger.warning(f"LLM fallback failed: {llm_err}")
                    
            return data

        except Exception as e:
            logger.error(f"PDF extraction failed for {file_path}: {e}")
            return {"error": str(e), "raw_text": ""}


# ── XLSX Extractor ───────────────────────────────────────────────────────────

def extract_xlsx(file_path: str) -> dict:
    """Extract structured data from an XLSX quotation."""
    with measure_latency("ingestion", "extract_xlsx", file=file_path):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            data = {
                "supplier_name": "",
                "doc_no": "",
                "doc_date": None,
                "valid_until": None,
                "project_name": "",
                "project_id": "",
                "currency": "INR",
                "payment_terms": "",
                "delivery_days": None,
                "line_items": [],
                "raw_text": "",
            }

            # Read all cells to build raw text and find structured fields
            all_text_parts = []
            for row in ws.iter_rows(values_only=False):
                row_texts = []
                for cell in row:
                    val = cell.value
                    if val is not None:
                        row_texts.append(str(val))
                all_text_parts.append(" | ".join(row_texts))

            raw_text = "\n".join(all_text_parts)
            data["raw_text"] = raw_text

            # Parse header info from first rows
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
                for cell in row:
                    val = str(cell.value or "").strip()
                    val_lower = val.lower()

                    # Look for supplier name (usually first merged cell)
                    if cell.row == 1 and cell.column == 1 and val and len(val) > 5:
                        data["supplier_name"] = val

                    # Look for quotation number
                    if "quotation no" in val_lower or val_lower == "quotation no:":
                        # Check next cell
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value:
                            data["doc_no"] = str(next_cell.value).strip()
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("SUP"):
                        if "-Q" in cell.value or "Q0" in cell.value:
                            data["doc_no"] = cell.value.strip()

                    # Look for date
                    if "date" in val_lower and ":" in val_lower:
                        date_part = val.split(":", 1)[-1].strip()
                        data["doc_date"] = _parse_date(date_part)
                    elif cell.row <= 5:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value and "date" in val_lower:
                            data["doc_date"] = _parse_date(str(next_cell.value))

                    # Project
                    if "project" in val_lower and ":" not in val_lower:
                        proj_match = re.search(r"Project:\s*(.+?)(?:\s*\(|$)", val)
                        if not proj_match:
                            proj_match = re.search(r"Project\s+(.+)", val)
                        if proj_match:
                            data["project_name"] = proj_match.group(1).strip()

            # Find line items (look for header row with Description/Unit/Qty)
            header_row = None
            desc_col = unit_col = qty_col = price_col = amt_col = None

            for row_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
                row_vals = [str(cell.value or "").lower() for cell in row]
                row_text = " ".join(row_vals)

                if "description" in row_text and ("qty" in row_text or "quantity" in row_text):
                    header_row = row_idx
                    for cell in row:
                        val = str(cell.value or "").lower()
                        if "desc" in val:
                            desc_col = cell.column
                        elif val in ("unit", "uom"):
                            unit_col = cell.column
                        elif "qty" in val or "quant" in val:
                            qty_col = cell.column
                        elif "unit price" in val or "rate" in val:
                            price_col = cell.column
                        elif "amount" in val or "total" in val:
                            amt_col = cell.column
                    break

            if header_row and desc_col:
                for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
                    desc_val = None
                    for cell in row:
                        if cell.column == desc_col:
                            desc_val = cell.value
                    if not desc_val or str(desc_val).strip() == "":
                        continue
                    if "sub-total" in str(desc_val).lower() or "total" in str(desc_val).lower():
                        break

                    item = {"description": str(desc_val).strip()}
                    for cell in row:
                        if unit_col and cell.column == unit_col:
                            item["unit"] = str(cell.value or "").strip()
                        elif qty_col and cell.column == qty_col:
                            item["quantity"] = _parse_float(cell.value)
                        elif price_col and cell.column == price_col:
                            item["unit_price"] = _parse_float(cell.value)
                        elif amt_col and cell.column == amt_col:
                            item["amount"] = _parse_float(cell.value)

                    if item.get("description"):
                        if not item.get("amount") and item.get("quantity") and item.get("unit_price"):
                            item["amount"] = round(item["quantity"] * item["unit_price"], 2)
                        data["line_items"].append(item)

            # Find terms section
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    val = str(cell.value or "")
                    val_lower = val.lower()
                    if "payment" in val_lower:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 2)
                        if next_cell.value:
                            data["payment_terms"] = str(next_cell.value).strip()
                    if "delivery" in val_lower:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 2)
                        if next_cell.value:
                            delivery_text = str(next_cell.value).strip()
                            days_match = re.search(r"(\d+)\s*days", delivery_text)
                            if days_match:
                                data["delivery_days"] = int(days_match.group(1))

            # LLM fallback if important fields are still missing
            _important_missing = (
                not data.get("line_items")
                or not data.get("supplier_name")
                or not data.get("doc_no")
                or not data.get("delivery_days")
            )
            if _important_missing and raw_text.strip():
                logger.info(f"Rule-based XLSX extraction incomplete, falling back to LLM for {file_path}")
                try:
                    from backend.intelligence.llm_client import extract_quotation_data
                    llm_data = extract_quotation_data(raw_text)
                    if llm_data:
                        for key, val in llm_data.items():
                            if key == "raw_text":
                                continue
                            existing = data.get(key)
                            if not existing or existing == 0:
                                data[key] = val
                except Exception as llm_err:
                    logger.warning(f"LLM fallback failed for XLSX: {llm_err}")

            return data

        except Exception as e:
            logger.error(f"XLSX extraction failed for {file_path}: {e}")
            return {"error": str(e), "raw_text": ""}


# ── CSV Extractor ────────────────────────────────────────────────────────────

def extract_csv(file_path: str) -> dict:
    """Extract structured data from a CSV quotation."""
    with measure_latency("ingestion", "extract_csv", file=file_path):
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()

            data = _extract_from_text(content, "csv")

            _important_missing = (
                not data.get("line_items")
                or not data.get("supplier_name")
            )
            if _important_missing:
                try:
                    from backend.intelligence.llm_client import extract_quotation_data
                    llm_data = extract_quotation_data(content)
                    if llm_data:
                        for key, val in llm_data.items():
                            if key == "raw_text":
                                continue
                            existing = data.get(key)
                            if not existing or existing == 0:
                                data[key] = val
                except Exception as llm_err:
                    logger.warning(f"LLM fallback failed for CSV: {llm_err}")

            return data

        except Exception as e:
            logger.error(f"CSV extraction failed for {file_path}: {e}")
            return {"error": str(e), "raw_text": ""}


# ── Email Extractor ──────────────────────────────────────────────────────────

def extract_email(file_path: str) -> dict:
    """Extract structured data from email text."""
    with measure_latency("ingestion", "extract_email", file=file_path):
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()

            data = _extract_from_text(content, "email")

            _important_missing = (
                not data.get("line_items")
                or not data.get("doc_no")
            )
            if _important_missing:
                try:
                    from backend.intelligence.llm_client import extract_quotation_data
                    llm_data = extract_quotation_data(content)
                    if llm_data:
                        for key, val in llm_data.items():
                            if key == "raw_text":
                                continue
                            existing = data.get(key)
                            if not existing or existing == 0:
                                data[key] = val
                except Exception as llm_err:
                    logger.warning(f"LLM fallback failed for email: {llm_err}")

            return data

        except Exception as e:
            logger.error(f"Email extraction failed for {file_path}: {e}")
            return {"error": str(e), "raw_text": ""}


# ── Scan Simulation Extractor ────────────────────────────────────────────────

def extract_scan(file_path: str) -> dict:
    """Extract structured data from scan simulation text."""
    with measure_latency("ingestion", "extract_scan", file=file_path):
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()

            data = _extract_from_text(content, "scan_sim")

            _important_missing = (
                not data.get("line_items")
                or not data.get("doc_no")
            )
            if _important_missing:
                try:
                    from backend.intelligence.llm_client import extract_quotation_data
                    llm_data = extract_quotation_data(content)
                    if llm_data:
                        for key, val in llm_data.items():
                            if key == "raw_text":
                                continue
                            existing = data.get(key)
                            if not existing or existing == 0:
                                data[key] = val
                except Exception as llm_err:
                    logger.warning(f"LLM fallback failed for scan: {llm_err}")

            return data

        except Exception as e:
            logger.error(f"Scan extraction failed for {file_path}: {e}")
            return {"error": str(e), "raw_text": ""}


# ── Shared Text Extraction Logic ─────────────────────────────────────────────

def _extract_from_text(text: str, source_type: str) -> dict:
    """Extract structured fields from raw text using regex patterns."""
    data = {
        "supplier_name": "",
        "doc_no": "",
        "doc_date": None,
        "valid_until": None,
        "validity_days": None,
        "project_name": "",
        "project_id": "",
        "currency": "INR",
        "payment_terms": "",
        "delivery_days": None,
        "warranty": "",
        "freight_terms": "",
        "total_incl_tax": None,
        "line_items": [],
        "raw_text": text,
    }

    lines = text.split("\n")

    # Extract supplier name (first non-empty line for scan/pdf)
    for line in lines[:5]:
        line = line.strip()
        if line and len(line) > 5 and not line.startswith(("From:", "To:", "Date:", "Subject:")):
            if source_type == "email":
                # For email, look for From: header
                from_match = re.search(r"From:\s*(.+?)(?:@|$)", text)
                if from_match:
                    # Supplier name might be in signature
                    pass
            else:
                data["supplier_name"] = line
            break

    # Extract quotation number
    quo_patterns = [
        r"(?:Quotation\s+No|Ref\s+No|Quote\s+Ref|QUOTATION)\s*[:\s]\s*(SUP\d+-\S+)",
        r"Ref\s+No\s*:\s*(\S+)",
        r"Quotation\s+No:\s*(\S+)",
    ]
    for pat in quo_patterns:
        match = re.search(pat, text, re.IGNORECASE)
        if match:
            data["doc_no"] = match.group(1).strip()
            break

    # Extract date
    date_patterns = [
        r"Date\s*:\s*(\d{1,2}[-/]\w{3}[-/]\d{4})",
        r"Date\s*:\s*(\d{1,2}[-/]\d{1,2}[-/]\d{4})",
        r"Date\s*:\s*(\d{4}-\d{2}-\d{2})",
        r"Date,(\d{4}-\d{2}-\d{2})",
    ]
    for pat in date_patterns:
        match = re.search(pat, text, re.IGNORECASE)
        if match:
            data["doc_date"] = _parse_date(match.group(1))
            break

    # Extract validity
    val_match = re.search(r"(?:Valid(?:ity)?|Valid\s+for|Valid\s+till)\s*[:\s]\s*(\d+)\s*days?",
                          text, re.IGNORECASE)
    if val_match:
        data["validity_days"] = int(val_match.group(1))

    # Extract project
    proj_patterns = [
        r"Project\s*[:\s]\s*(Project\s+\w+)",
        r"Project\s*[,:\s]\s*(\w+\s+\w+)",
        r"(PROJ-\w+)",
    ]
    for pat in proj_patterns:
        match = re.search(pat, text, re.IGNORECASE)
        if match:
            data["project_name"] = match.group(1).strip()
            break

    proj_id_match = re.search(r"(PROJ-\w+)", text)
    if proj_id_match:
        data["project_id"] = proj_id_match.group(1)

    # Extract currency
    if "USD" in text:
        data["currency"] = "USD"

    # Extract payment terms
    pay_match = re.search(r"Payment\s*(?:Terms?)?\s*[:\s]\s*(.+?)(?:\n|$)", text, re.IGNORECASE)
    if pay_match:
        data["payment_terms"] = pay_match.group(1).strip()

    # Extract delivery days — handles "18 days from receipt of PO & advance" and "Delivery: 18 days"
    del_match = re.search(
        r"Deliver(?:y)?\s*[:\s]\s*(\d+)\s*days?"
        r"|(\d+)\s*days?\s+from\s+(?:receipt\s+of\s+)?PO",
        text, re.IGNORECASE
    )
    if del_match:
        data["delivery_days"] = int(del_match.group(1) or del_match.group(2))

    # Extract warranty
    warranty_match = re.search(
        r"Warrant(?:y|ee)\s*[:\s]\s*(.+?)(?:\n|$)", text, re.IGNORECASE
    )
    if warranty_match:
        data["warranty"] = warranty_match.group(1).strip()

    # Extract freight terms
    freight_match = re.search(
        r"Freight\s*[:\s]\s*(.+?)(?:\n|$)", text, re.IGNORECASE
    )
    if freight_match:
        data["freight_terms"] = freight_match.group(1).strip()

    # Extract total incl tax (Grand Total / Total incl GST)
    total_incl_match = re.search(
        r"(?:GRAND\s+TOTAL|Total\s+(?:incl\.?\s*(?:GST|Tax)))\s*[:\s]*([₹\d,\.]+)",
        text, re.IGNORECASE
    )
    if total_incl_match:
        raw = total_incl_match.group(1).replace("₹", "").replace(",", "").strip()
        try:
            data["total_incl_tax"] = float(raw)
        except ValueError:
            pass

    # Extract line items (from CSV-format or general text)
    if source_type == "csv":
        _extract_csv_line_items(text, data)
    else:
        _extract_text_line_items(text, data)

    return data


def _extract_csv_line_items(text: str, data: dict):
    """Extract line items from CSV format text."""
    lines = text.split("\n")
    in_items = False
    for line in lines:
        if "SNo,Description" in line or "Description,Unit" in line:
            in_items = True
            continue
        if in_items and line.strip():
            parts = line.split(",")
            if len(parts) >= 5 and parts[0].strip().isdigit():
                try:
                    data["line_items"].append({
                        "description": parts[1].strip(),
                        "unit": parts[2].strip(),
                        "quantity": _parse_float(parts[3]),
                        "unit_price": _parse_float(parts[4]),
                        "amount": _parse_float(parts[5]) if len(parts) > 5 else 0,
                    })
                except (IndexError, ValueError):
                    pass
            elif "Sub-Total" in line or "Total" in line:
                break


def _extract_text_line_items(text: str, data: dict):
    """Extract line items from free-form text (PDF/email/scan)."""
    # Pattern for numbered items: "1. Description - Qty: 5 MT @ INR 68,000.00/MT = INR 340,000.00"
    item_pattern = re.compile(
        r"(\d+)\.\s+(.+?)\s*[-–]\s*"
        r"Qty:\s*([\d,.]+)\s*(\w+)\s*@\s*"
        r"(?:INR|USD|₹|\$)\s*([\d,.]+)/\w+\s*=\s*"
        r"(?:INR|USD|₹|\$)\s*([\d,.]+)",
        re.IGNORECASE
    )

    for match in item_pattern.finditer(text):
        data["line_items"].append({
            "description": match.group(2).strip(),
            "unit": match.group(4).strip(),
            "quantity": _parse_float(match.group(3)),
            "unit_price": _parse_float(match.group(5)),
            "amount": _parse_float(match.group(6)),
        })

    # If no items found, try tabular format (scan sim)
    if not data["line_items"]:
        tab_pattern = re.compile(
            r"(\d+)\s*\|\s*(.+?)\s*\|\s*([A-Za-z]+)\s*\|\s*([\d,.\w]+)\s*\|\s*([\d,.\w]+)\s*\|\s*([\d,.\w]+)"
        )
        for match in tab_pattern.finditer(text):
            data["line_items"].append({
                "description": match.group(2).strip(),
                "unit": match.group(3).strip(),
                "quantity": _parse_float(match.group(4)),
                "unit_price": _parse_float(match.group(5)),
                "amount": _parse_float(match.group(6)),
            })

    # If still no items, try space-separated tabular format (common in PDFs)
    if not data["line_items"]:
        space_tab_pattern = re.compile(
            r"^(\d+)\s+(.+?)\s+([A-Za-z]{2,4})\s+([\d,.\w]+)\s+([\d,.\w]+)\s+([\d,.\w]+)\s*$",
            re.MULTILINE
        )
        for match in space_tab_pattern.finditer(text):
            data["line_items"].append({
                "description": match.group(2).strip(),
                "unit": match.group(3).strip(),
                "quantity": _parse_float(match.group(4)),
                "unit_price": _parse_float(match.group(5)),
                "amount": _parse_float(match.group(6)),
            })
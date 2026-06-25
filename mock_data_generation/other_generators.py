"""
Generators for XLSX, CSV, email text, and scan-simulated image quotations.
"""
import random
import os
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

from mock_data_generation.data_definitions import (
    SUPPLIERS, PROJECTS, LINE_ITEMS_CATALOGUE, PAYMENT_TERMS,
    VALIDITY_DAYS_OPTIONS, get_doc_date, market_price, random_qty
)


# ─── XLSX Generator ───────────────────────────────────────────────────────────
def generate_xlsx(out_path, supplier, project, doc_no, doc_date,
                  line_items, validity_days, payment_terms,
                  missing_delivery=False, has_second_sheet=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    currency = supplier["currency"]

    # Column widths
    widths = {"A": 5, "B": 45, "C": 10, "D": 10, "E": 18, "F": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Styles
    hdr_fill = PatternFill("solid", fgColor="1A3A5C")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    title_font = Font(name="Arial", bold=True, size=13, color="1A3A5C")
    bold9 = Font(name="Arial", bold=True, size=9)
    reg9 = Font(name="Arial", size=9)
    total_fill = PatternFill("solid", fgColor="DDE8F5")
    alt_fill = PatternFill("solid", fgColor="F0F4F8")
    thin = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    row = 1

    # Supplier header block
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = supplier["name"]
    ws[f"A{row}"].font = title_font
    ws[f"A{row}"].alignment = left
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = supplier["address"]
    ws[f"A{row}"].font = reg9
    ws[f"E{row}"] = "Quotation No:"
    ws[f"E{row}"].font = bold9
    ws[f"F{row}"] = doc_no
    ws[f"F{row}"].font = reg9
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = f"GST: {supplier['gst'] or 'N/A'}   Tel: {supplier['phone']}"
    ws[f"A{row}"].font = reg9
    ws[f"E{row}"] = "Date:"
    ws[f"E{row}"].font = bold9
    ws[f"F{row}"] = doc_date.strftime("%d-%b-%Y")
    ws[f"F{row}"].font = reg9
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = f"Email: {supplier['email']}"
    ws[f"A{row}"].font = reg9
    ws[f"E{row}"] = "Valid Until:"
    ws[f"E{row}"].font = bold9
    ws[f"F{row}"] = (doc_date + timedelta(days=validity_days)).strftime("%d-%b-%Y")
    ws[f"F{row}"].font = reg9
    row += 1

    row += 1  # blank

    # Project info
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = f"Project: {project['name']}  ({project['id']})  —  {project['description']}"
    ws[f"A{row}"].font = bold9
    ws[f"A{row}"].fill = PatternFill("solid", fgColor="EDF2FA")
    row += 1

    row += 1  # blank

    # Column headers
    headers = ["#", "Description", "Unit", "Qty", f"Unit Price ({currency})", f"Amount ({currency})"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin_border
    ws.row_dimensions[row].height = 28
    row += 1

    data_start = row
    grand_total_col = "F"

    for idx, item in enumerate(line_items, 1):
        qty = item["qty"]
        up = item["unit_price"]
        amt_formula = f"=D{row}*E{row}"
        fill = alt_fill if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        vals = [idx, item["desc"], item["unit"], qty, up, amt_formula]
        aligns = [center, left, center, center, right, right]
        for ci, (v, al) in enumerate(zip(vals, aligns), 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = reg9
            cell.alignment = al
            cell.fill = fill
            cell.border = thin_border
            if ci in (5, 6):
                cell.number_format = '#,##0.00'
        row += 1

    data_end = row - 1

    row += 1  # blank

    # Totals
    subtotal_row = row
    ws[f"E{row}"] = "Sub-Total"
    ws[f"E{row}"].font = bold9
    ws[f"E{row}"].alignment = right
    ws[f"F{row}"] = f"=SUM(F{data_start}:F{data_end})"
    ws[f"F{row}"].number_format = '#,##0.00'
    ws[f"F{row}"].font = bold9
    ws[f"F{row}"].alignment = right
    row += 1

    if supplier["gst"] and currency == "INR":
        ws[f"E{row}"] = "GST @ 18%"
        ws[f"E{row}"].font = reg9
        ws[f"E{row}"].alignment = right
        ws[f"F{row}"] = f"=F{subtotal_row}*0.18"
        ws[f"F{row}"].number_format = '#,##0.00'
        ws[f"F{row}"].font = reg9
        ws[f"F{row}"].alignment = right
        gst_row = row
        row += 1

        ws[f"E{row}"] = "GRAND TOTAL"
        ws[f"E{row}"].font = Font(name="Arial", bold=True, size=10, color="1A3A5C")
        ws[f"E{row}"].alignment = right
        ws[f"F{row}"] = f"=F{subtotal_row}+F{gst_row}"
        ws[f"F{row}"].number_format = '#,##0.00'
        ws[f"F{row}"].font = Font(name="Arial", bold=True, size=10, color="1A3A5C")
        ws[f"F{row}"].fill = total_fill
        ws[f"F{row}"].alignment = right
    else:
        ws[f"E{row}"] = "TOTAL"
        ws[f"E{row}"].font = Font(name="Arial", bold=True, size=10, color="1A3A5C")
        ws[f"E{row}"].alignment = right
        ws[f"F{row}"] = f"=F{subtotal_row}"
        ws[f"F{row}"].number_format = '#,##0.00'
        ws[f"F{row}"].font = Font(name="Arial", bold=True, size=10, color="1A3A5C")
        ws[f"F{row}"].fill = total_fill
        ws[f"F{row}"].alignment = right

    row += 2

    # Terms
    terms = [
        ("Payment Terms:", payment_terms),
        ("Delivery:", f"{'To be confirmed' if missing_delivery else str(supplier['typical_lead_days'] + random.randint(-5,10)) + ' days from PO date'}"),
        ("Validity:", f"{validity_days} days"),
        ("Warranty:", "12 months from commissioning"),
        ("Freight:", "Inclusive" if currency == "INR" else "Ex-works, freight extra"),
    ]
    for label, val in terms:
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = bold9
        ws.merge_cells(f"C{row}:F{row}")
        ws[f"C{row}"] = val
        ws[f"C{row}"].font = reg9
        row += 1

    # Optional second sheet (item comparison / price history)
    if has_second_sheet:
        ws2 = wb.create_sheet("Price History")
        ws2["A1"] = "Previous Quotation Reference"
        ws2["A1"].font = bold9
        ws2["B1"] = doc_no.replace("R2", "R1")
        ws2["A2"] = "Note"
        ws2["B2"] = "Prices revised. See main sheet for updated rates."
        ws2["A2"].font = reg9
        ws2["B2"].font = reg9

    wb.save(out_path)


# ─── CSV Generator ────────────────────────────────────────────────────────────
def generate_csv(out_path, supplier, project, doc_no, doc_date,
                 line_items, validity_days, payment_terms):
    currency = supplier["currency"]
    lines = []
    lines.append(f"QUOTATION,{doc_no}")
    lines.append(f"Supplier,{supplier['name']}")
    lines.append(f"Contact,{supplier['contact']}")
    lines.append(f"Email,{supplier['email']}")
    lines.append(f"Date,{doc_date.strftime('%Y-%m-%d')}")
    lines.append(f"Valid Until,{(doc_date + timedelta(days=validity_days)).strftime('%Y-%m-%d')}")
    lines.append(f"Project,{project['name']}")
    lines.append(f"Project ID,{project['id']}")
    lines.append(f"Currency,{currency}")
    lines.append(f"Payment Terms,{payment_terms}")
    lines.append(f"Delivery Days,{supplier['typical_lead_days']}")
    lines.append("")
    lines.append("SNo,Description,Unit,Quantity,Unit Price,Amount")
    grand = 0
    for i, item in enumerate(line_items, 1):
        amt = round(item["qty"] * item["unit_price"], 2)
        grand += amt
        lines.append(f"{i},{item['desc']},{item['unit']},{item['qty']},{item['unit_price']},{amt}")
    lines.append(f",,,,,")
    lines.append(f",,,,Sub-Total,{grand:.2f}")
    if supplier["gst"] and currency == "INR":
        gst = round(grand * 0.18, 2)
        lines.append(f",,,,GST 18%,{gst:.2f}")
        lines.append(f",,,,Grand Total,{grand + gst:.2f}")
    else:
        lines.append(f",,,,Total,{grand:.2f}")

    with open(out_path, "w") as f:
        f.write("\n".join(lines))


# ─── Email Text Generator ─────────────────────────────────────────────────────
def generate_email_txt(out_path, supplier, project, doc_no, doc_date,
                        line_items, validity_days, payment_terms,
                        missing_delivery=False, vague_terms=False):
    currency = supplier["currency"]
    grand = sum(item["qty"] * item["unit_price"] for item in line_items)

    # Build a natural-sounding email (intentionally imperfect formatting)
    salutation = random.choice(["Dear Sir", "Dear Sir/Madam", "Hi", "Hello", "Respected Sir"])
    closings = ["Thanks & Regards", "Best Regards", "Warm Regards", "Regards", "Thanking you"]

    items_text = ""
    for i, item in enumerate(line_items, 1):
        amt = round(item["qty"] * item["unit_price"], 2)
        items_text += (f"  {i}. {item['desc']} - "
                       f"Qty: {item['qty']} {item['unit']} @ "
                       f"{currency} {item['unit_price']:,.2f}/{item['unit']} = {currency} {amt:,.2f}\n")

    delivery_line = ""
    if missing_delivery:
        delivery_line = "Delivery: Will confirm after receiving PO.\n"
    else:
        days = supplier["typical_lead_days"] + random.randint(-5, 10)
        delivery_line = f"Delivery: {days} days from receipt of PO and advance\n"

    vague_note = ""
    if vague_terms:
        vague_note = "\nNote: Other terms and conditions as mutually agreed.\n"

    if supplier["gst"] and currency == "INR":
        tax_line = f"GST @ 18%: INR {grand*0.18:,.2f}\nGrand Total: INR {grand*1.18:,.2f}"
    else:
        tax_line = f"Total: {currency} {grand:,.2f}"

    content = f"""From: {supplier['email']}
To: purchase@nexusolve.in
Date: {doc_date.strftime('%a, %d %b %Y')}
Subject: Quotation for {project['name']} - {project['category']} / Ref: {doc_no}

{salutation},

Thank you for the enquiry. Please find below our best offer for {project['name']}:

Quotation No: {doc_no}
Date: {doc_date.strftime('%d-%b-%Y')}
Valid for {validity_days} days

ITEMS QUOTED:
{items_text}
---
Sub-Total: {currency} {grand:,.2f}
{tax_line}
---

Payment Terms: {payment_terms}
{delivery_line}Warranty: 12 months from commissioning
{vague_note}
Please do the needful and release PO at earliest.

{random.choice(closings)},
{supplier['contact']}
{supplier['name']}
Tel: {supplier['phone']}
"""
    with open(out_path, "w") as f:
        f.write(content)


# ─── Scan Simulation (image-based PDF look) ───────────────────────────────────
def generate_scan_simulation_txt(out_path, supplier, project, doc_no, doc_date,
                                  line_items, validity_days, payment_terms):
    """
    Simulates what a scanned doc text looks like after OCR:
    slight mis-spacing, occasional OCR artefact chars, inconsistent capitalisation.
    Saved as .txt but named to indicate scan origin.
    """
    currency = supplier["currency"]
    grand = 0

    def ocr_noise(text):
        """Inject minor OCR-style noise."""
        replacements = {
            "0": random.choice(["0", "O", "0"]),
            "I": random.choice(["I", "l", "I"]),
            ",": random.choice([",", ".", ","]),
        }
        result = ""
        for ch in text:
            result += replacements.get(ch, ch)
        return result

    lines_out = []
    lines_out.append(f"{supplier['name'].upper()}")
    lines_out.append(f"{supplier['address']}")
    lines_out.append(f"Tel : {supplier['phone']}  |  {supplier['email']}")
    lines_out.append("")
    lines_out.append(f"QUOTATION")
    lines_out.append(f"Ref No  : {doc_no}")
    lines_out.append(f"Date    : {doc_date.strftime('%d-%m-%Y')}")
    lines_out.append(f"Valid   : {validity_days} Days")
    lines_out.append("")
    lines_out.append(f"To : Nexusolve Manufacturing Pvt Ltd")
    lines_out.append(f"Proj : {project['name']} / {project['id']}")
    lines_out.append("")
    lines_out.append(f"S.No  | {'Description':<40} | Unit | Qty  | Rate({currency})  | Amt({currency})")
    lines_out.append("-" * 90)

    for i, item in enumerate(line_items, 1):
        amt = round(item["qty"] * item["unit_price"], 2)
        grand += amt
        desc = item["desc"][:40]
        rate_str = ocr_noise(f"{item['unit_price']:,.2f}")
        amt_str = ocr_noise(f"{amt:,.2f}")
        lines_out.append(
            f"{i:<5} | {desc:<40} | {item['unit']:<4} | {str(item['qty']):<4} | {rate_str:>12} | {amt_str:>12}"
        )

    lines_out.append("-" * 90)
    if supplier["gst"] and currency == "INR":
        lines_out.append(f"{'':>60} SubTotal : {ocr_noise(f'{grand:,.2f}')}")
        lines_out.append(f"{'':>60} GST 18%  : {ocr_noise(f'{grand*0.18:,.2f}')}")
        lines_out.append(f"{'':>60} TOTAL    : {ocr_noise(f'{grand*1.18:,.2f}')}")
    else:
        lines_out.append(f"{'':>60} TOTAL    : {ocr_noise(f'{grand:,.2f}')}")
    lines_out.append("")
    lines_out.append(f"Payment : {payment_terms}")
    lines_out.append(f"Delivery : {supplier['typical_lead_days']} Days from PO")
    lines_out.append(f"Validity : {validity_days} Days")
    lines_out.append("")
    lines_out.append(f"For {supplier['name']}")
    lines_out.append("")
    lines_out.append(f"(Authorised Signatory)")
    lines_out.append(f"[STAMP]")  # simulates visible rubber stamp from scan

    with open(out_path, "w") as f:
        f.write("\n".join(lines_out))
